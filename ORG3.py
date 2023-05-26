import os
from datetime import datetime
from random import randint
from time import sleep as pause
from typing import Any, Iterable, Mapping, Optional

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side, PatternFill
from openpyxl.utils import get_column_letter
from selenium.webdriver import Chrome
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from tqdm import tqdm
from webdriver_manager.chrome import ChromeDriverManager

from pagination import process_page
#
pagen_list = process_page('https://ORG3.ru/')

start_time = datetime.now()


if not os.path.exists('resulting files'):
    os.makedirs('resulting files')


def get_date_and_time() -> str:
    return datetime.now().strftime('%d.%m.%y %H-%M-%S')


def to_excel(data: Iterable[Mapping[str, Any]], column_names: Iterable[str],
             file_name: Optional[str] = "table") -> None:
    """
    Создаёт из итерируемого объекта и имён столбцов
    xlsx файл в папке "resulting files".
    """
    wb = Workbook()
    worksheet = wb.active
    side = Side(border_style='thin')
    border = Border(
        left=side,
        right=side,
        top=side,
        bottom=side
    )
    alignment = Alignment(
        horizontal='left',
        vertical='center'
    )
    column_widths = []

    for column, name in enumerate(column_names, 1):
        cell = worksheet.cell(
            column=column,
            row=1,
            value=name
        )
        cell.font = Font(name='Calibri', size=11, bold=True)
        cell.fill = PatternFill('solid', fgColor='C5D9F1')
        cell.border = border
        cell.alignment = alignment
        column_widths.append(len(name) + 2)

    for row, product in enumerate(data, 2):
        if not product:
            print(row)
            continue
        for column, name in enumerate(column_names, 1):
            cell = worksheet.cell(
                column=column,
                row=row,
                value=product.get(name, '')
            )
            cell.font = Font(name='Calibri', size=11, bold=False)
            cell.border = border
            cell.alignment = alignment
            column_widths[column -
                          1] = max(column_widths[column -
                                                 1], len(str(cell.value)))

    for i, column_width in enumerate(column_widths, 1):
        worksheet.column_dimensions[get_column_letter(
            i)].width = column_width + 1

    datetime_now = get_date_and_time()
    wb.save(f"resulting files/{file_name} {datetime_now}.xlsx")

with Chrome(service=Service(ChromeDriverManager().install())) as driver:
    driver.maximize_window()

    print("Найдено страниц в каталоге:", len(pagen_list))

    pause(1)

    products = []
    column_names = [
        'Город',
        'Артикул',
        'Наименование',
        'Цена',
        'Примечание'
    ]

    # Проходим по всем страницам пагинации
    for i in pagen_list:
        url = i

        driver.get(url)
        pause(randint(6, 8))
        soup = BeautifulSoup(driver.page_source, 'lxml')


        # ищем все карточки с товаром на странице
        cards = soup.find('table', class_='products-list').find_all('tr', class_='products-list-item')
        city = soup.find('div', class_='header-office-address').text.split(',')[0].replace('г. ', '').strip()
        for card in cards:

            product = dict()

            try:
                WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.CLASS_NAME, 'item-final-price')))
            except Exception:
                pass

            product['Город'] = city
            product['Артикул'] = card.find('span', class_='product-item-code-copy').text
            product['Наименование'] = card.find('div', class_='products-list-item-name').text.strip()

            try:
                product['Примечание'] = card.find('div', class_='products-list-item-price-description').text
            except Exception:
                product['Примечание'] = ''


            try:
                product['Цена'] = card.find('div', class_='item-final-price').text
            except Exception:
                driver.refresh()
                WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.CLASS_NAME, 'item-final-price')))
                product['Цена'] = card.find('div', class_='item-final-price').text

                print(f'На странице {i} косяк с ценой')


            products.append(product)

for product in products:
    column_names.extend(list(set(product) - set(column_names)))

to_excel(products, column_names, file_name=f"elfgroup.ru_{city}")

print('Time taken:', str(datetime.now() - start_time).split('.')[0])
